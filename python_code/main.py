import customtkinter as ctk
import pyodbc
import datetime
import re 
from tkinter import ttk, messagebox 
from openpyxl.styles import Font
import os




class Database:
    def __init__(self):
        self.conn = self.get_db_connection()
        if not self.conn:
            messagebox.showerror("Error", "Database connection failed!")

    def get_db_connection(self):
        try:
            connection_string = (
                r'DRIVER={ODBC Driver 18 for SQL Server};'
                r'SERVER=.\SQLEXPRESS;'
                r'DATABASE=FinDB;'  # Changed to your database name
                r'TrustServerCertificate=yes;'
                r'Authentication=ActiveDirectoryIntegrated;'
            )
            conn = pyodbc.connect(connection_string)
            return conn
        except pyodbc.Error as ex:
            sqlstate = ex.args[0]
            print(f"Database connection error: {sqlstate}")
            messagebox.showerror("Error", f"Database connection error: {sqlstate}") # show DB error.
            return None
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {e}") # show generic error.
            return None

    def insert_revenue(self, revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description):
        """Inserts a new revenue record into the database."""
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = """
                INSERT INTO revenues (revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                params = (revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description)
                cursor.execute(query, params)
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}. Please ensure all field values are correct and try again.", icon='warning') # Combined message
                self.conn.rollback()
                return False
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                    cursor.close()
        return False

    def insert_expense(self, expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description):
        """Inserts a new expense record into the database."""
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = """
                INSERT INTO expenses (expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                params = (expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description)
                cursor.execute(query, params)
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}. Please ensure all field values are correct and try again.", icon='warning') # Combined message
                self.conn.rollback()
                return False
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                    cursor.close()
        return False



    def fetch_revenues(self):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                cursor.execute("SELECT revenue_id, revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description FROM revenues")
                return cursor.fetchall()
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                return []
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                    cursor.close()
        return []

    def fetch_expenses(self):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                cursor.execute("SELECT expense_id, expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description FROM expenses")
                return cursor.fetchall()
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                return []
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                    cursor.close()
        return []

    def update_expense(self, expense_id, expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = """
                UPDATE expenses
                SET expense_amount = ?, expense_type = ?, payment_method = ?, expense_source = ?, expense_date = ?, expense_description = ?
                WHERE expense_id = ?
                """
                params = (expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description, expense_id)
                cursor.execute(query, params)
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                self.conn.rollback()
                return False
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                  cursor.close()
        return False

    def delete_expense(self, expense_id):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = "DELETE FROM expenses WHERE expense_id = ?"
                cursor.execute(query, (expense_id,))
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                self.conn.rollback()
                return False
            finally:
              if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                cursor.close()
        return False

    def update_revenue(self, revenue_id, revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = """
                UPDATE revenues
                SET revenue_amount = ?, revenue_type = ?, payment_method = ?, income_source = ?, revenue_date = ?, revenue_description = ?
                WHERE revenue_id = ?
                """
                params = (revenue_amount, revenue_type, payment_method, income_source, revenue_date, revenue_description, revenue_id)
                cursor.execute(query, params)
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                self.conn.rollback()
                return False
            finally:
                if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                  cursor.close()
        return False

    def delete_revenue(self, revenue_id):
        if self.conn:
            try:
                cursor = self.conn.cursor()
                query = "DELETE FROM revenues WHERE revenue_id = ?"
                cursor.execute(query, (revenue_id,))
                self.conn.commit()
                return True
            except pyodbc.Error as ex:
                sqlstate = ex.args[0]
                print(f"Database error: {sqlstate}")
                messagebox.showerror("Error", f"Database error: {sqlstate}")
                self.conn.rollback()
                return False
            finally:
              if hasattr(locals(), 'cursor') and hasattr(cursor, 'close') and callable(cursor.close):
                cursor.close()
        return False


    def close(self):
        if self.conn:
            self.conn.close()





class FinApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Your Finance App!")
        self.geometry("800x700")

        self.db = Database()  # Instantiate the Database class

        # Initialize a dictionary to hold references to each Frame
        self.frames = {}

        # Main Frame
        self.main_frame = ctk.CTkFrame(self)
        self.frames["main"] = self.main_frame

        # Welcome Label (Pack it FIRST)
        self.label = ctk.CTkLabel(self.main_frame, text="Welcome!", font=ctk.CTkFont(size=24, weight="bold"))
        self.label.pack(pady=(50, 10))  # Reduced top padding

        # Total Revenue and Expense Labels Frame
        self.totals_frame = ctk.CTkFrame(self.main_frame)
        self.totals_frame.pack(pady=(20, 5), padx=20, fill="x") # Reduced top padding

        # Custom Font for Total Revenue and Expense Labels
        total_font = ctk.CTkFont(family="Arial", size=24, weight="bold")  # Example font

        self.total_revenue_label = ctk.CTkLabel(self.totals_frame, text="Total Revenue: $0.00", font=total_font)
        self.total_revenue_label.pack(side="top", anchor="w", fill="x", pady=10)  # Smaller padding

        self.total_expense_label = ctk.CTkLabel(self.totals_frame, text="Total Expense: $0.00", font=total_font)
        self.total_expense_label.pack(side="top", anchor="w", fill="x", pady=10)  # Smaller padding

        # Refresh Button (under Total Expense, centered)
        self.refresh_button = ctk.CTkButton(self.totals_frame, text="Refresh Values", command=self.update_total_labels, fg_color="#3b3b3b", text_color="white")
        self.refresh_button.pack(side="bottom", pady=5, anchor="center", fill="x")  # Centered below Expense

        # Financial Metrics Labels (Pack them THIRD - Grouped in a frame)
        self.financial_metrics_frame = ctk.CTkFrame(self.main_frame)  # Container for the metrics
        self.financial_metrics_frame.pack(pady=(120, 10), fill="x", padx=20, anchor="w")  # Reduced padding

        label_font = ctk.CTkFont(size=18, weight="normal")  # Larger font

        self.ratio_label = ctk.CTkLabel(self.financial_metrics_frame, text="Ratio: 0.00%", font=label_font, anchor="w")  # Left align
        self.ratio_label.pack(pady=(5, 2), anchor="w", fill="x", padx=20)  # Smaller padding

        self.profit_loss_label = ctk.CTkLabel(self.financial_metrics_frame, text="Net Profit/Loss: $0.00", font=label_font, anchor="w")
        self.profit_loss_label.pack(pady=(5, 2), anchor="w", fill="x", padx=20) # Smaller padding

        self.gross_profit_margin_label = ctk.CTkLabel(self.financial_metrics_frame, text="Gross Profit Margin: 0.00%", font=label_font, anchor="w")
        self.gross_profit_margin_label.pack(pady=(5, 2), anchor="w", fill="x", padx=20) # Smaller padding

        # Excel Export Buttons Frame
        self.excel_buttons_frame = ctk.CTkFrame(self.main_frame)
        self.excel_buttons_frame.pack(pady=(10, 20), fill="x", padx=20)

        # Excel Export Revenue Button
        self.excel_export_revenue_button = ctk.CTkButton(self.excel_buttons_frame, text="Export Revenue to Excel", command=self.export_revenue_to_excel, fg_color="#396e17", text_color="white", corner_radius=8)
        self.excel_export_revenue_button.pack(side="top", fill='x', pady=5)  # Separate row

        # Excel Export Expense Button
        self.excel_export_expense_button = ctk.CTkButton(self.excel_buttons_frame, text="Export Expenses to Excel", command=self.export_expense_to_excel, fg_color="#396e17", text_color="white", corner_radius=8)
        self.excel_export_expense_button.pack(side="top", fill='x', pady=5)  # Separate row

        # Button Frame
        self.button_frame = ctk.CTkFrame(self)
        self.button_frame.pack(side="bottom", fill="x", padx=20, pady=20)

        self.expense_button = ctk.CTkButton(self.button_frame, text="Expense Management", command=lambda: self.show_frame("expense"))
        self.expense_button.pack(side="left", fill="x", expand=True, padx=10)

        self.revenue_button = ctk.CTkButton(self.button_frame, text="Revenue Management", command=lambda: self.show_frame("revenue"))
        self.revenue_button.pack(side="right", fill="x", expand=True, padx=10)

        # Expense Frame (Initially Hidden)
        self.expense_frame = ExpenseFrame(self, self.db)
        self.frames["expense"] = self.expense_frame

        # Expense Detail Frame (Initially Hidden)
        self.expense_detail_frame = ExpenseDetailFrame(self, self.db)
        self.frames["expense_detail"] = self.expense_detail_frame

        # Revenue Frame (Initially Hidden)
        self.revenue_frame = RevenueFrame(self, self.db)
        self.frames["revenue"] = self.revenue_frame

        # Revenue Detail Frame (Initially Hidden)
        self.revenue_detail_frame = RevenueDetailFrame(self, self.db)
        self.frames["revenue_detail"] = self.revenue_detail_frame

        # Pack the initial frame
        self.show_frame("main")

    def show_frame(self, frame_name):
        """Hides all frames and shows the specified frame."""
        for name, frame in self.frames.items():
            frame.pack_forget()
        self.frames[frame_name].pack(fill="both", expand=True, padx=20, pady=20)
        if frame_name == 'main':
            self.update_total_labels()  # show value when back to main frame

    def update_total_labels(self):
        """Updates all financial labels with data from the database."""
        total_revenue = self.calculate_total_revenue()
        total_expense = self.calculate_total_expense()

        # Calculate the ratio
        if total_expense == 0:
            ratio = 0.0  # Avoid division by zero
        else:
            ratio = (total_revenue / total_expense) * 100

        # Calculate net profit/loss
        net_profit_loss = total_revenue - total_expense

        # Calculate gross profit margin
        if total_revenue == 0:
            gross_profit_margin = 0.0  # Avoid division by zero
        else:
            gross_profit_margin = ((total_revenue - total_expense) / total_revenue) * 100


        self.total_revenue_label.configure(text=f"Total Revenue: ${total_revenue:.2f}")
        self.total_expense_label.configure(text=f"Total Expense: ${total_expense:.2f}")
        self.ratio_label.configure(text=f"Ratio: {ratio:.2f}%")
        self.profit_loss_label.configure(text=f"Net Profit/Loss: ${net_profit_loss:.2f}")
        self.gross_profit_margin_label.configure(text=f"Gross Profit Margin: {gross_profit_margin:.2f}%")

    def export_revenue_to_excel(self):
        """Exports revenue data to an Excel file."""
        try:
            revenues = self.db.fetch_revenues()
            if not revenues:
                messagebox.showinfo("Info", "No revenue data to export.")
                return

            # Create a new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Revenue Data"

            # Define header row with bold font
            header_font = Font(bold=True)
            header = ["Revenue ID", "Amount", "Type", "Payment Method", "Income Source", "Date", "Description"]
            for col_num, column_title in enumerate(header, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = header_font

            # Write data rows
            for row_num, revenue in enumerate(revenues, 2):  # Start from row 2
                for col_num, cell_value in enumerate(revenue, 1):
                    sheet.cell(row=row_num, column=col_num).value = cell_value

            # Save the workbook to a file
            file_path = "revenue_data.xlsx"
            workbook.save(file_path)
            messagebox.showinfo("Success", f"Revenue data exported to {file_path}")

            # Open the file if possible
            if os.path.exists(file_path):
                os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting revenue to Excel: {e}")

    def export_expense_to_excel(self):
        """Exports expense data to an Excel file."""
        try:
            expenses = self.db.fetch_expenses()
            if not expenses:
                messagebox.showinfo("Info", "No expense data to export.")
                return

            # Create a new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Expense Data"

            # Define header row with bold font
            header_font = Font(bold=True)
            header = ["Expense ID", "Amount", "Type", "Payment Method", "Expense Source", "Date", "Description"]
            for col_num, column_title in enumerate(header, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = header_font

            # Write data rows
            for row_num, expense in enumerate(expenses, 2):  # Start from row 2
                for col_num, cell_value in enumerate(expense, 1):
                    sheet.cell(row=row_num, column=col_num).value = cell_value

            # Save the workbook to a file
            file_path = "expense_data.xlsx"
            workbook.save(file_path)
            messagebox.showinfo("Success", f"Expense data exported to {file_path}")

            # Open the file if possible
            if os.path.exists(file_path):
                os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting expense to Excel: {e}")


    def calculate_total_revenue(self):
        """Calculates the total revenue from the database."""
        try:
            revenues = self.db.fetch_revenues()
            total = sum(revenue[1] for revenue in revenues)  # Assuming revenue_amount is at index 1
            return total
        except Exception as e:
            messagebox.showerror("Error", f"Error calculating total revenue: {e}")
            return 0

    def calculate_total_expense(self):
        """Calculates the total expenses from the database."""
        try:
            expenses = self.db.fetch_expenses()
            total = sum(expense[1] for expense in expenses)  # Assuming expense_amount is at index 1
            return total
        except Exception as e:
            messagebox.showerror("Error", f"Error calculating total expenses: {e}")
            return 0
    def show_expense_detail_frame(self):
        self.show_frame("expense_detail")

    def show_expense_frame(self):  # Add this method
        self.show_frame("expense")

    def show_revenue_detail_frame(self):
        self.show_frame("revenue_detail")

    def show_revenue_frame(self):
        self.show_frame("revenue")



###########################
# Expence
###########################

class ExpenseFrame(ctk.CTkFrame):
    def __init__(self, master, db):  # Take 'master' and 'db' as arguments
        super().__init__(master)
        self.master = master # Store the master (FinApp instance)
        self.db = db # Store the database instance

        self.label = ctk.CTkLabel(self, text="Expense Management", font=ctk.CTkFont(size=20, weight="bold"))
        self.label.pack(pady=20)

        # Expense Detail Button (Top Right)
        self.expense_detail_button = ctk.CTkButton(
            self,
            text="Expense Detail",
            command=self.show_expense_detail,
            fg_color="#444444",  # Gray Background
            text_color="white"      # White Text
        )
        self.expense_detail_button.place(relx=0.98, rely=0.02, anchor="ne") # Place at top right

        # Back to Main Button (Top Left)
        self.back_button = ctk.CTkButton(
            self,
            text="← Back to Main",
            command=self.go_back_to_main,
            fg_color="#d0d51e",  # Green Background
            text_color="black"   # Black Text
        )
        self.back_button.place(relx=0.02, rely=0.02, anchor="nw") # Place at top left

        # Expense Amount
        self.expense_amount_label = ctk.CTkLabel(self, text="Expense Amount:")
        self.expense_amount_label.pack(pady=5)
        self.expense_amount_entry = ctk.CTkEntry(self, placeholder_text="Expense Amount")
        self.expense_amount_entry.pack(pady=5)

        # Expense Type
        self.expense_type_label = ctk.CTkLabel(self, text="Expense Type:")
        self.expense_type_label.pack(pady=5)
        self.expense_type_options = ['Rent', 'Utilities', 'Salary', 'Supplies', 'Marketing', 'Other']
        self.expense_type_combobox = ctk.CTkComboBox(self, values=self.expense_type_options)
        self.expense_type_combobox.pack(pady=5)

        # Payment Method
        self.payment_method_label = ctk.CTkLabel(self, text="Payment Method:")
        self.payment_method_label.pack(pady=5)
        self.payment_method_options = ['Cash', 'Credit Card', 'Debit Card', 'Bank Transfer', 'PayPal', 'Else']
        self.payment_method_combobox = ctk.CTkComboBox(self, values=self.payment_method_options)
        self.payment_method_combobox.pack(pady=5)

        # Expense Source
        self.expense_source_label = ctk.CTkLabel(self, text="Expense Source:")
        self.expense_source_label.pack(pady=5)
        self.expense_source_options = ['Checking Account', 'Savings Account', 'Credit Card', 'Other']
        self.expense_source_combobox = ctk.CTkComboBox(self, values=self.expense_source_options)
        self.expense_source_combobox.pack(pady=5)

        # Expense Date
        self.expense_date_label = ctk.CTkLabel(self, text="Expense Date:") # Note the format
        self.expense_date_label.pack(pady=5)
        self.expense_date_entry = ctk.CTkEntry(self, placeholder_text="YYYY-MM-DD")
        self.expense_date_entry.pack(pady=5)

        # Expense Description
        self.expense_description_label = ctk.CTkLabel(self, text="Expense Description:")
        self.expense_description_label.pack(pady=5)
        self.expense_description_textbox = ctk.CTkTextbox(self, height=50)
        self.expense_description_textbox.pack(pady=5)

        # Add Expense Button
        self.add_expense_button = ctk.CTkButton(self, text="Add Expense", command=self.add_expense)
        self.add_expense_button.pack(pady=10)


    def show_expense_detail(self):
        self.master.show_expense_detail_frame()

    def go_back_to_main(self):
        self.pack_forget()
        self.master.main_frame.pack(fill="both", expand=True, padx=20, pady=20) # Use master to access main_frame

    def add_expense(self):
        expense_amount = self.expense_amount_entry.get()
        expense_type = self.expense_type_combobox.get()
        payment_method = self.payment_method_combobox.get()
        expense_source = self.expense_source_combobox.get()
        expense_date = self.expense_date_entry.get()
        expense_description = self.expense_description_textbox.get("0.0", "end") # Get all text from textbox

        # Validate that the fields are not empty
        if not all([expense_amount, expense_type, payment_method, expense_source, expense_date]):
          messagebox.showerror("Error", "All fields are required.", icon='warning')
          return

        # Validate date format using regular expression
        date_pattern = r"^\d{4}-\d{2}-\d{2}$"
        if not re.match(date_pattern, expense_date):
          messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.", icon='warning')
          return

        try:
            expense_amount = float(expense_amount)  # Convert after validation
            # Call the database function to insert the expense
            success = self.db.insert_expense(expense_amount, expense_type, payment_method, expense_source, expense_date, expense_description)
            if not success:
              return # already handled
            messagebox.showinfo("Success", "Expense added successfully!")
            # Clear the input fields after successful insertion
            self.expense_amount_entry.delete(0, 'end')
            self.expense_date_entry.delete(0, 'end')
            self.expense_description_textbox.delete("0.0", 'end')  # Clear the textbox as well

        except ValueError:
            messagebox.showerror("Error", "Invalid amount. Please enter a number.", icon='warning')
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')







class ExpenseDetailFrame(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master)
        self.master = master
        self.db = db

        self.label = ctk.CTkLabel(self, text="Expense Detail", font=ctk.CTkFont(size=20, weight="bold"))
        self.label.pack(pady=10)

        self.back_button = ctk.CTkButton(
            self,
            text="Back to Expense",
            command=self.go_back_to_expense,
            fg_color="#444444",  # Gray background
            text_color="white"   # White text
        )
        self.back_button.place(relx=0.98, rely=0.02, anchor="ne")  # Top right corner


        # Treeview setup
        self.tree = ttk.Treeview(
            self,
            columns=(
                "expense_id",
                "expense_amount",
                "expense_type",
                "payment_method",
                "expense_source",
                "expense_date",
                "expense_description",
            ),
            show="headings",
        )

        self.tree.heading("expense_id", text="ID")
        self.tree.heading("expense_amount", text="Amount")
        self.tree.heading("expense_type", text="Type")
        self.tree.heading("payment_method", text="Payment Method")
        self.tree.heading("expense_source", text="Source")
        self.tree.heading("expense_date", text="Date")
        self.tree.heading("expense_description", text="Description")

        self.tree.column("expense_id", width=50, anchor="center")
        self.tree.column("expense_amount", width=80, anchor="center")
        self.tree.column("expense_type", width=100, anchor="center")
        self.tree.column("payment_method", width=100, anchor="center")  # Corrected line
        self.tree.column("expense_source", width=100, anchor="center")
        self.tree.column("expense_date", width=80, anchor="center")
        self.tree.column("expense_description", width=200, anchor="center")

        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("Treeview.Heading", background="#343638", foreground="white", font=('Arial', 10, 'bold'),padding=10)
        style.configure("Treeview", background="#2B2B2F", foreground="white", fieldbackground="#2B2B2F", font=('Arial', 9))
        style.map("Treeview", background=[("selected", "#565B5E")])
        style.map("Treeview.Heading", background=[('active', '#444444')])

        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.vsb.set)

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.vsb.pack(side="right", fill="y")

        # CRUD Buttons Frame
        self.crud_frame = ctk.CTkFrame(self)
        self.crud_frame.pack(pady=5)

        self.add_button = ctk.CTkButton(self.crud_frame, text="Add", command=self.add_expense, fg_color="#3cff3c", text_color="black")
        self.add_button.pack(side="left", padx=5)

        self.update_button = ctk.CTkButton(self.crud_frame, text="Update", command=self.update_expense, fg_color="#3c9bff", text_color="black")
        self.update_button.pack(side="left", padx=5)

        self.delete_button = ctk.CTkButton(self.crud_frame, text="Delete", command=self.delete_expense, fg_color="#ff3c3c", text_color="black")
        self.delete_button.pack(side="left", padx=5)

        self.clear_button = ctk.CTkButton(self.crud_frame, text="Clear", command=self.clear_selection, fg_color="#ffff3c", text_color="black")
        self.clear_button.pack(side="left", padx=5)

        # Entry fields for adding/updating expenses
        self.entry_frame = ctk.CTkFrame(self)
        self.entry_frame.pack(pady=5)

        self.amount_label = ctk.CTkLabel(self.entry_frame, text="Amount:")
        self.amount_label.grid(row=0, column=0, padx=5, pady=2)
        self.amount_entry = ctk.CTkEntry(self.entry_frame)
        self.amount_entry.grid(row=0, column=1, padx=5, pady=2)

        self.type_label = ctk.CTkLabel(self.entry_frame, text="Type:")
        self.type_label.grid(row=0, column=2, padx=5, pady=2)
        expense_types = ['Rent', 'Utilities', 'Salary', 'Supplies', 'Marketing', 'Other']
        self.type_combobox = ctk.CTkComboBox(self.entry_frame, values=expense_types)
        self.type_combobox.grid(row=0, column=3, padx=5, pady=2)

        self.payment_label = ctk.CTkLabel(self.entry_frame, text="Payment:")
        self.payment_label.grid(row=0, column=4, padx=5, pady=2)
        payment_methods = ['Cash', 'Credit Card', 'Debit Card', 'Bank Transfer', 'PayPal', 'Else']
        self.payment_combobox = ctk.CTkComboBox(self.entry_frame, values=payment_methods)
        self.payment_combobox.grid(row=0, column=5, padx=5, pady=2)

        self.source_label = ctk.CTkLabel(self.entry_frame, text="Source:")
        self.source_label.grid(row=1, column=0, padx=5, pady=2)
        expense_sources = ['Checking Account', 'Savings Account', 'Credit Card', 'Other']
        self.source_combobox = ctk.CTkComboBox(self.entry_frame, values=expense_sources)
        self.source_combobox.grid(row=1, column=1, padx=5, pady=2)

        self.date_label = ctk.CTkLabel(self.entry_frame, text="Date (YYYY-MM-DD):")
        self.date_label.grid(row=1, column=2, padx=5, pady=2)
        self.date_entry = ctk.CTkEntry(self.entry_frame)
        self.date_entry.grid(row=1, column=3, padx=5, pady=2)

        self.description_label = ctk.CTkLabel(self.entry_frame, text="Description:")
        self.description_label.grid(row=1, column=4, padx=5, pady=2)
        self.description_entry = ctk.CTkEntry(self.entry_frame)
        self.description_entry.grid(row=1, column=5, padx=5, pady=2)


        self.load_expenses()
        self.tree.bind("<ButtonRelease-1>", self.select_item) # Binding

    def go_back_to_expense(self):
        self.pack_forget()
        self.master.show_expense_frame()

    def load_expenses(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        expenses = self.db.fetch_expenses()

        if expenses:
            for expense in expenses:
                # Format the date and amount for better readability
                formatted_expense = list(expense)  # Convert to list for modification
                formatted_expense[1] = f"{expense[1]:.2f}"  # Format amount to 2 decimal places
                if isinstance(expense[5], datetime.date):  # Check if the date is a datetime.date object
                    formatted_expense[5] = expense[5].strftime('%Y-%m-%d')  # Format the date to YYYY-MM-DD
                self.tree.insert("", "end", values=formatted_expense)
        else:
            ctk.CTkLabel(self, text="No expenses found in the database.", text_color="red").pack(pady=10)


    def add_expense(self):
        try:
            amount = float(self.amount_entry.get())
            expense_type = self.type_combobox.get()
            payment_method = self.payment_combobox.get()
            source = self.source_combobox.get()
            date_str = self.date_entry.get()
            description = self.description_entry.get()

            # Convert date string to datetime.date object
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

            success = self.db.insert_expense(amount, expense_type, payment_method, source, date, description)

            if success:
                messagebox.showinfo("Success", "Expense added successfully!")
                self.load_expenses()  # Refresh the Treeview
                self.clear_entries()
            else:
                messagebox.showerror("Error", "Failed to add expense.")

        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    def update_expense(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select an expense to update.")
            return

        try:
            expense_id = self.tree.item(selected_item, 'values')[0]
            print(f"Attempting to update expense with ID: {expense_id}")  # Debug print
            amount = float(self.amount_entry.get())
            expense_type = self.type_combobox.get()
            payment_method = self.payment_combobox.get()
            source = self.source_combobox.get()
            date_str = self.date_entry.get()
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            description = self.description_entry.get()


            # Assuming you have an 'update_expense' method in your Database class:
            success = self.db.update_expense(expense_id, amount, expense_type, payment_method, source, date, description)
            print(f"Update successful: {success}") # Debug print

            if success:
                messagebox.showinfo("Success", "Expense updated successfully!")
                self.load_expenses() # Refresh the Treeview
                self.clear_entries()

            else:
                messagebox.showerror("Error", "Failed to update expense.")

        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
        except Exception as e:
          messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')


    def delete_expense(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select an expense to delete.")
            return

        try:
            expense_id = self.tree.item(selected_item, 'values')[0]
            print(f"Attempting to delete expense with ID: {expense_id}")  # Debug print
            # Confirm before delete
            if messagebox.askyesno("Confirm", "Are you sure you want to delete this expense?"):
                success = self.db.delete_expense(expense_id)
                print(f"Delete successful: {success}")  # Debug print

                if success:
                    messagebox.showinfo("Success", "Expense deleted successfully!")
                    self.load_expenses()  # Refresh the Treeview
                    self.clear_entries()
                else:
                    messagebox.showerror("Error", "Failed to delete expense.")

        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')

    def clear_selection(self):
        self.tree.selection_remove(self.tree.selection()) # Deselect any selected items
        self.clear_entries()

    def clear_entries(self):
        self.amount_entry.delete(0, 'end')
        self.date_entry.delete(0, 'end')
        self.description_entry.delete(0, 'end')

    def select_item(self, event):
        selected_item = self.tree.selection()
        if selected_item:
            # Clear the entry fields before populating them
            self.clear_entries()

            values = self.tree.item(selected_item, 'values')
            self.amount_entry.insert(0, values[1])  # Amount
            self.type_combobox.set(values[2])      # Type
            self.payment_combobox.set(values[3])   # Payment
            self.source_combobox.set(values[4])    # Source
            self.date_entry.insert(0, values[5])    # Date
            self.description_entry.insert(0, values[6]) # Description


















###########################
# Revenue
###########################

class RevenueFrame(ctk.CTkFrame):
    def __init__(self, master, db):  # Take 'master' and 'db' as arguments
        super().__init__(master)
        self.master = master  # Store the main (FinApp instance)
        self.db = db  # Store the database instance

        self.label = ctk.CTkLabel(self, text="Revenue Management", font=ctk.CTkFont(size=20, weight="bold"))
        self.label.pack(pady=20)

        # Revenue Detail Button (Top Right)
        self.revenue_detail_button = ctk.CTkButton(
            self,
            text="Revenue Detail",
            command=self.show_revenue_detail,
            fg_color="#444444",  # Gray Background
            text_color="white"  # White Text
        )
        self.revenue_detail_button.place(relx=0.98, rely=0.02, anchor="ne")  # Place at top right

        # Back to Main Button (Top Left)
        self.back_button = ctk.CTkButton(
            self,
            text="← Back to Main",
            command=self.go_back_to_main,
            fg_color="#d0d51e",  # Green Background
            text_color="black"  # Black Text
        )
        self.back_button.place(relx=0.02, rely=0.02, anchor="nw")  # Place at top left

        # Revenue Amount
        self.revenue_amount_label = ctk.CTkLabel(self, text="Revenue Amount:")
        self.revenue_amount_label.pack(pady=5)
        self.revenue_amount_entry = ctk.CTkEntry(self, placeholder_text="Revenue Amount")
        self.revenue_amount_entry.pack(pady=5)

        # Revenue Type
        self.revenue_type_label = ctk.CTkLabel(self, text="Revenue Type:")
        self.revenue_type_label.pack(pady=5)
        self.revenue_type_options = ['Sales', 'Service Fee', 'Interest', 'Donation', 'Rent', 'Other']
        self.revenue_type_combobox = ctk.CTkComboBox(self, values=self.revenue_type_options)
        self.revenue_type_combobox.pack(pady=5)

        # Payment Method
        self.payment_method_label = ctk.CTkLabel(self, text="Payment Method:")
        self.payment_method_label.pack(pady=5)
        self.payment_method_options = ['Cash', 'Credit Card', 'Debit Card', 'Bank Transfer', 'PayPal', 'Else']
        self.payment_method_combobox = ctk.CTkComboBox(self, values=self.payment_method_options)
        self.payment_method_combobox.pack(pady=5)

        # Income Source
        self.income_source_label = ctk.CTkLabel(self, text="Income Source:")
        self.income_source_label.pack(pady=5)
        self.income_source_options = ['Main Business', 'Side Hustle', 'Investment', 'Other']
        self.income_source_combobox = ctk.CTkComboBox(self, values=self.income_source_options)
        self.income_source_combobox.pack(pady=5)

        # Revenue Date
        self.revenue_date_label = ctk.CTkLabel(self, text="Revenue Date:")  # Note the format
        self.revenue_date_label.pack(pady=5)
        self.revenue_date_entry = ctk.CTkEntry(self, placeholder_text="YYYY-MM-DD")
        self.revenue_date_entry.pack(pady=5)

        # Revenue Description
        self.revenue_description_label = ctk.CTkLabel(self, text="Revenue Description:")
        self.revenue_description_label.pack(pady=5)
        self.revenue_description_textbox = ctk.CTkTextbox(self, height=50)
        self.revenue_description_textbox.pack(pady=5)

        # Add Revenue Button
        self.add_revenue_button = ctk.CTkButton(self, text="Add Revenue", command=self.add_revenue)
        self.add_revenue_button.pack(pady=10)

    def show_revenue_detail(self):
        self.master.show_revenue_detail_frame()

    def go_back_to_main(self):
        self.pack_forget()
        self.master.main_frame.pack(fill="both", expand=True, padx=20, pady=20)  # Use master to access main_frame

    def add_revenue(self):
        revenue_amount = self.revenue_amount_entry.get()
        revenue_type = self.revenue_type_combobox.get()
        payment_method = self.payment_method_combobox.get()
        income_source = self.income_source_combobox.get()
        revenue_date = self.revenue_date_entry.get()
        revenue_description = self.revenue_description_textbox.get("0.0", "end")  # Get all text from textbox

        # Validate that the fields are not empty
        if not all([revenue_amount, revenue_type, payment_method, income_source, revenue_date]):
            messagebox.showerror("Error", "All fields are required.", icon='warning')
            return

        # Validate date format using regular expression
        date_pattern = r"^\d{4}-\d{2}-\d{2}$"
        if not re.match(date_pattern, revenue_date):
            messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.", icon='warning')
            return

        try:
            revenue_amount = float(revenue_amount)  # Convert after validation
            # Call the database function to insert the revenue
            success = self.db.insert_revenue(revenue_amount, revenue_type, payment_method, income_source, revenue_date,
                                             revenue_description)
            if not success:
                return  # already handled
            messagebox.showinfo("Success", "Revenue added successfully!")
            # Clear the input fields after successful insertion
            self.revenue_amount_entry.delete(0, 'end')
            self.revenue_date_entry.delete(0, 'end')
            self.revenue_description_textbox.delete("0.0", 'end')  # Clear the textbox as well

        except ValueError:
            messagebox.showerror("Error", "Invalid amount. Please enter a number.", icon='warning')
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')


class RevenueDetailFrame(ctk.CTkFrame):

    def __init__(self, master, db):
        super().__init__(master)
        self.master = master
        self.db = db

        self.label = ctk.CTkLabel(self, text="Revenue Detail", font=ctk.CTkFont(size=20, weight="bold"))
        self.label.pack(pady=10)

        self.back_button = ctk.CTkButton(
            self,
            text="Back to Revenue",
            command=self.go_back_to_revenue,
            fg_color="#444444",  # Gray background
            text_color="white"  # White text
        )
        self.back_button.place(relx=0.98, rely=0.02, anchor="ne")  # Top right corner

        # Treeview setup
        self.tree = ttk.Treeview(
            self,
            columns=(
                "revenue_id",
                "revenue_amount",
                "revenue_type",
                "payment_method",
                "income_source",
                "revenue_date",
                "revenue_description",
            ),
            show="headings",
        )

        self.tree.heading("revenue_id", text="ID")
        self.tree.heading("revenue_amount", text="Amount")
        self.tree.heading("revenue_type", text="Type")
        self.tree.heading("payment_method", text="Payment Method")
        self.tree.heading("income_source", text="Source")
        self.tree.heading("revenue_date", text="Date")
        self.tree.heading("revenue_description", text="Description")

        self.tree.column("revenue_id", width=50, anchor="center")
        self.tree.column("revenue_amount", width=80, anchor="center")
        self.tree.column("revenue_type", width=100, anchor="center")
        self.tree.column("payment_method", width=100, anchor="center")  # Corrected line
        self.tree.column("income_source", width=100, anchor="center")
        self.tree.column("revenue_date", width=80, anchor="center")
        self.tree.column("revenue_description", width=200, anchor="center")

        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("Treeview.Heading", background="#343638", foreground="white", font=('Arial', 10, 'bold'),
                        padding=10)
        style.configure("Treeview", background="#2B2B2F", foreground="white", fieldbackground="#2B2B2F",
                        font=('Arial', 9))
        style.map("Treeview", background=[("selected", "#565B5E")])
        style.map("Treeview.Heading", background=[('active', '#444444')])

        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.vsb.set)

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.vsb.pack(side="right", fill="y")

        # CRUD Buttons Frame
        self.crud_frame = ctk.CTkFrame(self)
        self.crud_frame.pack(pady=5)

        self.add_button = ctk.CTkButton(self.crud_frame, text="Add", command=self.add_revenue, fg_color="#3cff3c",
                                         text_color="black")
        self.add_button.pack(side="left", padx=5)

        self.update_button = ctk.CTkButton(self.crud_frame, text="Update", command=self.update_revenue,
                                            fg_color="#3c9bff", text_color="black")
        self.update_button.pack(side="left", padx=5)

        self.delete_button = ctk.CTkButton(self.crud_frame, text="Delete", command=self.delete_revenue,
                                            fg_color="#ff3c3c", text_color="black")
        self.delete_button.pack(side="left", padx=5)

        self.clear_button = ctk.CTkButton(self.crud_frame, text="Clear", command=self.clear_selection,
                                           fg_color="#ffff3c", text_color="black")
        self.clear_button.pack(side="left", padx=5)

        # Entry fields for adding/updating revenues
        self.entry_frame = ctk.CTkFrame(self)
        self.entry_frame.pack(pady=5)

        self.amount_label = ctk.CTkLabel(self.entry_frame, text="Amount:")
        self.amount_label.grid(row=0, column=0, padx=5, pady=2)
        self.amount_entry = ctk.CTkEntry(self.entry_frame)
        self.amount_entry.grid(row=0, column=1, padx=5, pady=2)

        self.type_label = ctk.CTkLabel(self.entry_frame, text="Type:")
        self.type_label.grid(row=0, column=2, padx=5, pady=2)
        revenue_types = ['Sales', 'Service Fee', 'Interest', 'Donation', 'Rent', 'Other']
        self.type_combobox = ctk.CTkComboBox(self.entry_frame, values=revenue_types)
        self.type_combobox.grid(row=0, column=3, padx=5, pady=2)

        self.payment_label = ctk.CTkLabel(self.entry_frame, text="Payment:")
        self.payment_label.grid(row=0, column=4, padx=5, pady=2)
        payment_methods = ['Cash', 'Credit Card', 'Debit Card', 'Bank Transfer', 'PayPal', 'Else']
        self.payment_combobox = ctk.CTkComboBox(self.entry_frame, values=payment_methods)
        self.payment_combobox.grid(row=0, column=5, padx=5, pady=2)

        self.source_label = ctk.CTkLabel(self.entry_frame, text="Source:")
        self.source_label.grid(row=1, column=0, padx=5, pady=2)
        income_sources = ['Main Business', 'Side Hustle', 'Investment', 'Other']
        self.source_combobox = ctk.CTkComboBox(self.entry_frame, values=income_sources)
        self.source_combobox.grid(row=1, column=1, padx=5, pady=2)

        self.date_label = ctk.CTkLabel(self.entry_frame, text="Date (YYYY-MM-DD):")
        self.date_label.grid(row=1, column=2, padx=5, pady=2)
        self.date_entry = ctk.CTkEntry(self.entry_frame)
        self.date_entry.grid(row=1, column=3, padx=5, pady=2)

        self.description_label = ctk.CTkLabel(self.entry_frame, text="Description:")
        self.description_label.grid(row=1, column=4, padx=5, pady=2)
        self.description_entry = ctk.CTkEntry(self.entry_frame)
        self.description_entry.grid(row=1, column=5, padx=5, pady=2)

        self.load_revenues()
        self.tree.bind("<ButtonRelease-1>", self.select_item)  # Binding

    def go_back_to_revenue(self):
        self.pack_forget()
        self.master.show_revenue_frame()

    def load_revenues(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        revenues = self.db.fetch_revenues()

        if revenues:
            for revenue in revenues:
                # Format the date and amount for better readability
                formatted_revenue = list(revenue)  # Convert to list for modification
                formatted_revenue[1] = f"{revenue[1]:.2f}"  # Format amount to 2 decimal places
                if isinstance(revenue[5], datetime.date):  # Check if the date is a datetime.date object
                    formatted_revenue[5] = revenue[5].strftime('%Y-%m-%d')  # Format the date to YYYY-MM-DD
                self.tree.insert("", "end", values=formatted_revenue)
        else:
            ctk.CTkLabel(self, text="No revenues found in the database.", text_color="red").pack(pady=10)

    def add_revenue(self):
        try:
            amount = float(self.amount_entry.get())
            revenue_type = self.type_combobox.get()
            payment_method = self.payment_combobox.get()
            source = self.source_combobox.get()
            date_str = self.date_entry.get()
            description = self.description_entry.get()

            # Convert date string to datetime.date object
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

            success = self.db.insert_revenue(amount, revenue_type, payment_method, source, date, description)

            if success:
                messagebox.showinfo("Success", "Revenue added successfully!")
                self.load_revenues()  # Refresh the Treeview
                self.clear_entries()
            else:
                messagebox.showerror("Error", "Failed to add revenue.")

        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    def update_revenue(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a revenue to update.")
            return

        try:
            revenue_id = self.tree.item(selected_item, 'values')[0]
            print(f"Attempting to update revenue with ID: {revenue_id}")  # Debug print
            amount = float(self.amount_entry.get())
            revenue_type = self.type_combobox.get()
            payment_method = self.payment_combobox.get()
            source = self.source_combobox.get()
            date_str = self.date_entry.get()
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            description = self.description_entry.get()


            success = self.db.update_revenue(revenue_id, amount, revenue_type, payment_method, source, date,
                                             description)
            print(f"Update successful: {success}")  # Debug print

            if success:
                messagebox.showinfo("Success", "Revenue updated successfully!")
                self.load_revenues()  # Refresh the Treeview
                self.clear_entries()

            else:
                messagebox.showerror("Error", "Failed to update revenue.")

        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')

    def delete_revenue(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a revenue to delete.")
            return

        try:
            revenue_id = self.tree.item(selected_item, 'values')[0]
            print(f"Attempting to delete revenue with ID: {revenue_id}")  # Debug print
            # Confirm before delete
            if messagebox.askyesno("Confirm", "Are you sure you want to delete this revenue?"):
                success = self.db.delete_revenue(revenue_id)
                print(f"Delete successful: {success}")  # Debug print

                if success:
                    messagebox.showinfo("Success", "Revenue deleted successfully!")
                    self.load_revenues()  # Refresh the Treeview
                    self.clear_entries()
                else:
                    messagebox.showerror("Error", "Failed to delete revenue.")

        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}", icon='warning')

    def clear_selection(self):
        self.tree.selection_remove(self.tree.selection())  # Deselect any selected items
        self.clear_entries()

    def clear_entries(self):
        self.amount_entry.delete(0, 'end')
        self.date_entry.delete(0, 'end')
        self.description_entry.delete(0, 'end')

    def select_item(self, event):


        selected_item = self.tree.selection()
        if selected_item:
            # Clear the entry fields before populating them
            self.clear_entries()

            values = self.tree.item(selected_item, 'values')
            self.amount_entry.insert(0, values[1])  # Amount
            self.type_combobox.set(values[2])  # Type
            self.payment_combobox.set(values[3])  # Payment
            self.source_combobox.set(values[4])  # Source
            self.date_entry.insert(0, values[5])  # Date
            self.description_entry.insert(0, values[6])  # Description







if __name__ == "__main__":
    app = FinApp()
    app.mainloop()
